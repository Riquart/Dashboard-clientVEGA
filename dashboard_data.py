"""
Extraction des données du fichier de suivi clientèle (Excel) vers un dict JSON.

La feuille "Données" est organisée en blocs par profession. Chaque bloc commence
par une ligne d'en-tête (colonne A = nom de la profession, colonnes suivantes = dates),
puis ~42 lignes d'indicateurs.

On extrait, en granularité hebdomadaire (médiane par semaine ISO, robuste aux relevés
journaliers corrompus), les indicateurs de base et d'activité (connectés sur 30 jours
glissants), pour abonnés / non-abonnés, PS et licences.
"""

import datetime
import statistics

import openpyxl

# Ligne d'en-tête de chaque bloc dans la feuille "Données"
BLOCKS = {
    "Infirmier": 1,
    "Kiné": 44,
    "Orthophoniste": 87,
    "Orthoptiste": 130,
    "Pédicure-Podologue": 173,
    "Total": 215,
}
PROFS = ["Infirmier", "Kiné", "Orthophoniste", "Orthoptiste", "Pédicure-Podologue"]

# Bornes : on ignore les dates aberrantes (< 2015) et les colonnes futures vides
LOWER = datetime.datetime(2015, 1, 1)
CUTOFF = datetime.datetime(2026, 5, 25)  # ajuster si le fichier va plus loin

METRICS = [
    "ps_ab_base", "ps_na_base", "lic_ab_base", "lic_na_base",
    "ps_ab_act", "ps_na_act", "lic_ab_act", "lic_na_act",
]


def _block_ranges():
    hrs = list(BLOCKS.values())
    ranges = {}
    for i, (name, hr) in enumerate(BLOCKS.items()):
        end = (hrs[i + 1] - 2) if i + 1 < len(hrs) else hr + 41
        ranges[name] = (hr, end)
    return ranges


def _find_row(ws, hr, end, pred):
    for r in range(hr + 1, end + 1):
        a = ws.cell(row=r, column=1).value
        if a and pred(str(a).strip()):
            return r
    return None


def _rows_for(ws, hr, end):
    return {
        "ps_ab_base": _find_row(ws, hr, end, lambda a: a.startswith("Ps - Abonnés")),
        "ps_na_base": _find_row(ws, hr, end, lambda a: a.startswith("Ps - Non Abonnés")),
        "lic_ab_base": _find_row(ws, hr, end, lambda a: a.startswith("Licences - Abonnées") and "moyenne" not in a and "delta" not in a),
        "lic_na_base": _find_row(ws, hr, end, lambda a: a.startswith("Licences - Non Abonnées")),
        "ps_ab_act": _find_row(ws, hr, end, lambda a: a.startswith("PS connectés mois glissant - Abonnés")),
        "ps_na_act": _find_row(ws, hr, end, lambda a: a.startswith("PS connectés mois glissant - Non abonnés")),
        "lic_ab_act": _find_row(ws, hr, end, lambda a: a.startswith("Licences connectées mois glissant - Abonnés")),
        "lic_na_act": _find_row(ws, hr, end, lambda a: a.startswith("Licences connectées mois glissant - Non abonnés")),
    }


def _num(ws, r, c):
    if r is None:
        return None
    v = ws.cell(row=r, column=c).value
    return float(v) if isinstance(v, (int, float)) else None


def extract(xlsx_path, sheet="Données"):
    """Lit le fichier Excel et renvoie un dict {profession: {labels, <metrics>}}."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet]
    ranges = _block_ranges()
    data = {}

    for name, (hr, end) in ranges.items():
        rmap = _rows_for(ws, hr, end)
        weeks = {}  # (iso_year, iso_week) -> {monday, metric: [values]}
        for c in range(2, ws.max_column + 1):
            dt = ws.cell(row=hr, column=c).value
            if not (isinstance(dt, datetime.datetime) and LOWER <= dt <= CUTOFF):
                continue
            iso = dt.isocalendar()
            key = (iso[0], iso[1])
            monday = dt - datetime.timedelta(days=dt.weekday())
            wk = weeks.setdefault(key, {"monday": monday, **{m: [] for m in METRICS}})
            wk["monday"] = min(wk["monday"], monday)
            for m in METRICS:
                v = _num(ws, rmap[m], c)
                if v is not None and v >= 0:
                    wk[m].append(v)

        series = {"labels": []}
        for m in METRICS:
            series[m] = []
        for key in sorted(weeks):
            wk = weeks[key]
            if not wk["ps_ab_base"]:
                continue
            series["labels"].append(wk["monday"].strftime("%Y-%m-%d"))
            for m in METRICS:
                series[m].append(round(statistics.median(wk[m])) if wk[m] else None)
        data[name] = series

    wb.close()

    # Cohérence : Total = somme des 5 professions quand toutes présentes
    L = data["Total"]["labels"]
    for m in METRICS:
        for i in range(len(L)):
            vals = [data[p][m][i] for p in PROFS
                    if i < len(data[p][m]) and data[p][m][i] is not None]
            if len(vals) == len(PROFS):
                data["Total"][m][i] = round(sum(vals))

    return data


if __name__ == "__main__":
    import json
    import sys
    path = sys.argv[1] if len(sys.argv) > 1 else "data/source.xlsx"
    d = extract(path)
    L = d["Total"]["labels"]
    print(f"Semaines : {len(L)} ({L[0]} -> {L[-1]})")
    n = -1
    s = d["Total"]
    print(f"Total — non-abonnés actifs : {s['ps_na_act'][n]} / base {s['ps_na_base'][n]}")
    print(json.dumps(d, ensure_ascii=False)[:200], "...")
